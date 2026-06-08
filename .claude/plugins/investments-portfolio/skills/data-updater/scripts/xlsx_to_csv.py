#!/usr/bin/env python3
"""미래에셋 퇴직연금 상품제안서 xlsx -> resource/*.csv 변환기.

기존 resource/*.csv 와 동일한 산출물을 생성한다 (LibreOffice "표시된 값으로 저장"):
- 입력: '실적배당형(펀드, ETF)' 시트
- 숫자 서식 적용 (0.00 -> 소수 2자리, #,##0 -> 천단위 콤마)
- 25 컬럼, UTF-8 BOM, CRLF (update_fund_data.py 가 소비하는 형식)

검증: 26년03월 xlsx -> csv 변환 결과가 커밋된 26년03월 CSV 와 byte 단위로 일치함.

의존성: openpyxl (update_fund_data.py 와 달리 이 스크립트만 외부 패키지 필요).

Usage:
    python xlsx_to_csv.py --src <xlsx_path> --dst <csv_path>
"""
import argparse
import csv
import datetime as dt
from decimal import Decimal, ROUND_HALF_UP, Context

import openpyxl

SHEET = "실적배당형(펀드, ETF)"
NCOLS = 25  # 커밋된 resource CSV 는 모두 25 컬럼

# Excel/LibreOffice 는 숫자를 15 유효자리로 먼저 정규화한 뒤 셀 서식으로 반올림한다.
# (예: 이진수 0.20499999999999996 -> "0.205" -> 0.21). 이를 재현해야 기존 CSV 와 일치.
_CTX15 = Context(prec=15, rounding=ROUND_HALF_UP)


def _round(value, decimals):
    d15 = _CTX15.create_decimal(Decimal(repr(value)))
    q = Decimal(1).scaleb(-decimals)
    d = d15.quantize(q, rounding=ROUND_HALF_UP)
    if d == 0:
        d = abs(d)  # -0.00 -> 0.00
    return d


def fmt_cell(value, number_format):
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, dt.datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, (int, float)):
        nf = (number_format or "General").split(";")[0]  # drop [Red] negative part
        if nf == "General":
            if isinstance(value, int) or float(value).is_integer():
                return str(int(value))
            return repr(value)
        if nf == "0":
            return str(int(_round(value, 0)))
        if nf == "0.00":
            return f"{_round(value, 2):.2f}"
        if nf.startswith("#,##0"):
            if "0.00" in nf:
                return f"{_round(value, 2):,.2f}"
            return f"{int(_round(value, 0)):,}"
        if isinstance(value, int) or float(value).is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def convert(src, dst):
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET}' not found in {src}. Sheets: {wb.sheetnames}")
    ws = wb[SHEET]
    out_rows = []
    max_used = 0
    for row in ws.iter_rows():
        cells = [fmt_cell(c.value, c.number_format) for c in row]
        for i, v in enumerate(cells):
            if v != "":
                max_used = max(max_used, i)
        cells = cells[:NCOLS]
        if len(cells) < NCOLS:
            cells += [""] * (NCOLS - len(cells))
        out_rows.append(cells)
    if max_used >= NCOLS:
        raise SystemExit(
            f"Data found in column index {max_used} (>= {NCOLS}); would be truncated."
        )
    with open(dst, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f, lineterminator="\r\n").writerows(out_rows)
    return len(out_rows)


def main():
    ap = argparse.ArgumentParser(description="Convert 퇴직연금 상품제안서 xlsx -> resource CSV")
    ap.add_argument("--src", required=True, help="입력 xlsx 경로")
    ap.add_argument("--dst", required=True, help="출력 csv 경로")
    args = ap.parse_args()
    n = convert(args.src, args.dst)
    print(f"Wrote {n} rows -> {args.dst}")


if __name__ == "__main__":
    main()
